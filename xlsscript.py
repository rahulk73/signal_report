from openpyxl import Workbook,load_workbook
from openpyxl.worksheet.table import Table,TableStyleInfo
from openpyxl.styles import Alignment,Font,PatternFill
import openpyxl.utils.cell as cell
import datetime
import re
from os import path as os_path
from os import mkdir
from sqlscript import GetSignalDataAnalog,GetSignalDataEvent,AccessDeniedError
import pickle
class SingleSheet():
    def __init__(self,all_signal_data,object_fullpaths,signal_type,data_type,customization,template_path,file_path,column=3):
        self.all_signal_data = all_signal_data
        self.object_fullpaths = object_fullpaths
        self.signal_type = signal_type
        self.data_type = data_type
        self.customization = customization
        self.template_path = template_path
        self.file_path = file_path
        self.column = column
        self.sheet_index = 1
        self.chart_counter = 0
        self.changes_mv_fields = (
            'Value Meantype',
            'Value Object UID',
            'Value Quality',
            'Measurement',
        )
        self.consumption_mt_fields = (
            'Value Meantype',
            'Value Object UID',
            'Value Quality',
            'Metering',
            'Consumption',
        )
        self.average_mv_fields = (
            'Value Meantype',
            'Value Object UID',
            'Value Quality',
            'Measurement',
            'Average',
            'Max',
            'Min',
        )
        self.changes_event_fields = (
            'Event Mess',
            'Event Userlooked',
        )
        self.font_head =  Font(b=True,color="4EB1BA",sz=15)
        self.alt_font_head = Font(b=True,color='FFFFFF',sz=15)
        self.align_head = Alignment(horizontal='center',vertical='center')
        self.fill_head = PatternFill('solid',fgColor='222930')
        self.alt_fill_head = PatternFill('solid',fgColor='FF8362')
        self.style = TableStyleInfo(name='TableStyleMedium9',showColumnStripes=True,showRowStripes=False)
        
        if self.template_path:
            self.wb = load_workbook(filename=self.template_path)
            self.ws_template = self.wb.active
            self.ws = self.wb.copy_worksheet(self.ws_template)
            self.wb.remove(self.ws_template)
            self.sheet_index+=1
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
        self.ws.title = 'Report'+str(self.sheet_index)
        self.cur_row = self.ws.max_row+2
        self.not_found=[]
        self.build_all()
        self.wb.save(self.file_path)
        self.wb.close()
    def build_all(self):
        self.build_data()
        self.build_table()
    def build_data(self):
        self.table_type = {
            'All Measurements': {
                'Changes':4,
                'Average':7,
            },
            'All Metering': {
                'Changes':4,
                'Consumption':5
            },
        }
        self.zipped_data = []
        for index,signal_data in enumerate(self.all_signal_data):
            if signal_data and self.zipped_data == []:
                self.zipped_data = signal_data
            elif signal_data and self.zipped_data:
                self.zipped_data = [i+j[self.column-1:] for i,j in zip(self.zipped_data,signal_data)]
            elif not signal_data:
                self.not_found.append(self.object_fullpaths.pop(index))
        self.step = self.table_type.get(self.signal_type,dict()).get(self.data_type,None)
    def build_table(self):
        column=self.column
        for i,path in enumerate(self.object_fullpaths,start=1):
            for j in range(self.step):
                self.ws.column_dimensions[cell.get_column_letter(column+j)].width=25
            base_cell = self.ws.cell(row=self.cur_row, column=column)
            self.ws.merge_cells(start_row=self.cur_row,end_row=self.cur_row+2,start_column=column,end_column=column+self.step-1)
            base_cell.value = path
            base_cell.alignment = self.align_head
            if i%2:            
                base_cell.font = self.alt_font_head
                base_cell.fill = self.alt_fill_head
            else:
                base_cell.font = self.font_head
                base_cell.fill = self.fill_head
            if not self.customization['Excel'][1]:
                self.ws.column_dimensions.group(cell.get_column_letter(column),cell.get_column_letter(column+self.step-1),hidden=False,outline_level=i)
            column+=self.step
        self.ws.cell(row=self.cur_row+3,column=1,value='Date')
        self.ws.cell(row=self.cur_row+3,column=2,value='Time')
        self.ws.column_dimensions['A'].width=15
        self.ws.column_dimensions['B'].width=15

        column = self.column
        fields = {4:self.changes_mv_fields,5:self.consumption_mt_fields,7:self.average_mv_fields}.get(self.step)
        for i,field in enumerate([[*fields]]*len(self.object_fullpaths),start=1):
            for j in range(self.step):
                self.ws.cell(row=self.cur_row+3,column=column+j,value=field[j]+'-'+str(i))
            column+=self.step

        for row in self.zipped_data:
            self.ws.append(row)

        table_row_start = self.cur_row+3
        table_column_end = (self.column-1)+len(self.object_fullpaths)*self.step
        table_row_end = len(self.zipped_data)+table_row_start

        tab = Table(displayName='Report',ref="A"+str(table_row_start)+":"+cell.get_column_letter(table_column_end)+str(table_row_end))
        tab.tableStyleInfo = self.style
        self.ws.add_table(tab)
        if self.customization['Excel'][3] and self.signal_type in ['All Measurements','All Metering']:
            for i in range(6,6+self.step*(len(self.object_fullpaths)-1)+1,self.step):
                if self.data_type in ['Average','Consumption']:
                    self.build_chart(min_col=i,max_col=i+1,min_row=table_row_start,max_row=table_row_end)
                else:
                    self.build_chart(min_col=i,min_row=table_row_start,max_row=table_row_end)
    def build_chart(self,min_col,min_row,max_row,max_col=None):
        def type1():
            from openpyxl.chart import AreaChart
            chart = AreaChart()
            return chart
        def type2():
            from openpyxl.chart import BarChart
            chart = BarChart()
            chart.type='col'
            return chart
        def type3():
            from openpyxl.chart import BarChart
            chart = BarChart()
            chart.type='bar'
            return chart
        def type4():
            from openpyxl.chart import LineChart
            chart = LineChart()
            return chart
        def type0():
            return
        self.chart_options = {
            1:type1,
            2:type2,
            3:type3,
            4:type4,
            0:type0,
        }
        chart = self.chart_options[self.customization['Excel'][3]]()
        from openpyxl.chart import Reference
        chart.style = 42
        chart.title = self.object_fullpaths[self.chart_counter]
        self.chart_counter+=1
        data = Reference(self.ws,min_col=min_col,max_col=max_col,min_row=min_row,max_row=max_row)
        chart.add_data(data,titles_from_data=True)
        cats = Reference(self.ws,min_col=1,max_col=2,min_row=min_row+1,max_row=max_row)
        chart.set_categories(cats)
        self.ws.add_chart(chart,cell.get_column_letter(min_col-3)+str(self.ws.max_row+1))
class MultipleSheet(SingleSheet):
    def __init__(self,*args):
        super(MultipleSheet,self).__init__(*args,column=1)
    def build_all(self):
        self.build_data()
        for index,signal_data in enumerate(self.all_signal_data):
            if signal_data:
                self.build_table(index)
            else:
                self.not_found.append(self.object_fullpaths[index])
    def build_data(self):
        self.table_type = {
            'All Measurements': {
                'Changes':6,
                'Average':9,
            },
            'All Metering': {
                'Changes':6,
                'Consumption':7,
            },
        }
        self.step=self.table_type.get(self.signal_type,dict()).get(self.data_type,4)
    def build_table(self,index):
        column = self.column
        for j in range(1,self.step+1):
            self.ws.column_dimensions[cell.get_column_letter(j)].width=25
        base_cell = self.ws.cell(row=self.cur_row, column=column)
        self.ws.merge_cells(start_row=self.cur_row,end_row=self.cur_row+2,start_column=column,end_column=column+self.step-1)
        base_cell.value = self.object_fullpaths[index]
        base_cell.alignment = self.align_head
        if index%2:            
            base_cell.font = self.alt_font_head
            base_cell.fill = self.alt_fill_head
        else:
            base_cell.font = self.font_head
            base_cell.fill = self.fill_head
        fields = {4:self.changes_event_fields,6:self.changes_mv_fields,7:self.consumption_mt_fields,9:self.average_mv_fields}.get(self.step)
        for i,head in enumerate(['Date','Time',*fields]):
            self.ws.cell(row=self.cur_row+3,column=column+i,value=head)
        for row in self.all_signal_data[index]:
            self.ws.append(row)

        table_row_start = self.cur_row+3
        table_column_end = self.step
        table_row_end = len(self.all_signal_data[index])+table_row_start

        tab = Table(displayName='Report'+str(index),ref="A"+str(table_row_start)+":"+cell.get_column_letter(table_column_end)+str(table_row_end))
        tab.tableStyleInfo = self.style
        self.ws.add_table(tab)
        if self.signal_type in ['All Measurements','All Metering']:
            if self.data_type in ['Average','Consumption']:
                self.build_chart(min_col=6,max_col=7,min_row=table_row_start,max_row=table_row_end)
            else:
                self.build_chart(min_col=self.step,min_row=table_row_start,max_row=table_row_end)
        if self.template_path:
            self.ws = self.wb.copy_worksheet(self.ws_template)
            self.ws.title = 'Report-'+str(self.sheet_index)
            self.sheet_index+=1
        else:
            self.ws = self.wb.create_sheet()
        self.cur_row = self.ws.max_row+2 
class SingleSheetEvent(SingleSheet):
    def __init__(self,*args):
        super(SingleSheetEvent,self).__init__(*args,column=1)
    def build_data(self):
        self.changes_all_event_fields = (
            'Origin',
            'Description',
            'Message',
            'Quality',
            'Command Origin',

        )
        self.step = 7
    def build_table(self):
        column = self.column
        for j in range(1,self.step+1):
            self.ws.column_dimensions[cell.get_column_letter(j)].width=25
        self.ws.column_dimensions['C'].width=self.ws.column_dimensions['D'].width=60
        base_cell = self.ws.cell(row=self.cur_row, column=column)
        self.ws.merge_cells(start_row=self.cur_row,end_row=self.cur_row+2,start_column=column,end_column=column+self.step-1)
        base_cell.value = 'Event Report'
        base_cell.alignment = self.align_head
        base_cell.font = self.font_head
        base_cell.fill = self.fill_head
        for i,head in enumerate(['Date','Time',*self.changes_all_event_fields]):
            self.ws.cell(row=self.cur_row+3,column=column+i,value=head) 
        for row in self.all_signal_data:
            self.ws.append(row)
        table_row_start = self.cur_row+3
        table_column_end = self.step
        table_row_end = len(self.all_signal_data)+table_row_start

        tab = Table(displayName='Report',ref="A"+str(table_row_start)+":"+cell.get_column_letter(table_column_end)+str(table_row_end))
        tab.tableStyleInfo = self.style
        self.ws.add_table(tab)

class Preferences:
    options_default = {
    'Excel':{
        1:0,
        2:0,
        3:4,
    },
    'Template':{
        1:'',
    },
    'MySQL':{
        'sh':'localhost',
        'un':'mcisadmin',
        'pw':'s$e!P!C!L@2014',
        'db':'pacis',
    }
}
class Increment():
    def __init__(self,vv):
        if(isinstance(vv,float)):
            self.cur=vv
            self.inc=[0]
        else:
            self.inc=[vv]
            self.cur=None
    def next(self,vv):
        if(isinstance(vv,float)):
            if self.cur:
                self.inc.append(vv-self.cur)
                self.cur=vv
            else:
                self.inc.append(0)
                self.cur=vv                                    
        else:
            self.inc.append(vv)

class ExcessiveDataError(Exception):
    pass
class AnalogReport:
    def __init__(self,fdate,tdate,timezone,hidden,object_fullpaths,signal_type,fq=None,radb=None,data_type='Changes',template_path=None):
        try:
            self.str_time=datetime.datetime.now().strftime('%y%m%d%H%M%S')
            if not os_path.isdir('./SignalLog'):
                mkdir('./SignalLog')
            self.file_path = './SignalLog/'+self.str_time+' Analog.xlsx'
            self.template_path = template_path #= "C:/Users/OISM/Desktop/SignalReportApp/Templates/template1.xlsx"
            self.fdate = fdate
            self.tdate = tdate
            self.timezonediff = (-1)*int(re.findall('GMT(.+):00',timezone)[0])
            self.fdate += datetime.timedelta(hours=self.timezonediff)
            self.tdate += datetime.timedelta(hours=self.timezonediff)
            self.timezonediff*=-1
            self.object_fullpaths = list(object_fullpaths)
            self.signal_type = signal_type
            self.fq = fq
            self.radb = radb
            self.data_type = data_type
            if os_path.isfile('settings'):
                with open('settings','rb') as file:
                    self.customization = pickle.load(file)
            else:
                self.customization = Preferences.options_default
            self.signal_info=[]
            self.all_signal_data=[]
            for path in self.object_fullpaths:
                self.signal_info.append(GetSignalDataAnalog(
                    path,
                    self.signal_type,
                    self.fdate,
                    self.tdate,
                    **self.customization['MySQL']
                ))
            for signal in self.signal_info:
                self.data=[]
                self.unparsed_data = signal.result
                self.uid = signal.uid
                if not self.unparsed_data:
                    self.all_signal_data.append([])
                    continue
                elif signal_type in ['All Measurements','All Metering']:
                    if not hidden:
                        self.analog_advanced()
                    else:
                        self.analog_raw()
                else:
                    self.event_selected()
                self.all_signal_data.append(self.data)
            if not self.isListEmpty(self.all_signal_data):
                if (not hidden and not self.customization['Excel'][2]):
                    self.workbook = SingleSheet(
                        self.all_signal_data,
                        self.object_fullpaths,
                        self.signal_type,
                        self.data_type,
                        self.customization,
                        self.template_path,
                        self.file_path,
                    )
                else:
                    self.workbook = MultipleSheet(
                        self.all_signal_data,
                        self.object_fullpaths,
                        self.signal_type,
                        self.data_type,
                        self.customization,
                        self.template_path,
                        self.file_path,

                    )
                if self.workbook.not_found:
                    self.result = 2
                else:
                    self.result = 1
            else:
                self.result = 0
        except AccessDeniedError:
            self.result = -1
        except PermissionError:
            self.result = -2
        except ExcessiveDataError:
            self.result = -3
        except Exception as e:
            self.result = -4
            self.errormessage = str(e)
            print(e)
    def analog_advanced(self):
        self.option = {
            1:1.0,
            2:60.0,
            3:float(60*60),
            4:float(24*60*60),
            5:float(7*24*60*60),
            6:float(30*7*24*60*60)
        }
        self.roundTo=(float)(self.fq)*self.option.get(self.radb)
        self.found_first={self.uid:0}
        next_dt = self.fdate
        cur_val=()
        for vm,vo,dt,ms,vq,vv in self.unparsed_data:
            dt+=datetime.timedelta(milliseconds=ms)
            if len(self.data)>12800:
                raise ExcessiveDataError
            elif dt>self.tdate:
                break
            elif dt>next_dt:
                if self.found_first[vo] == 0:
                    if cur_val:
                        self.incrementdata = Increment(cur_val[3])
                        self.average_aux = [
                            float(cur_val[3]),
                            1.0,
                            float(cur_val[3]),
                            float(cur_val[3])
                        ]
                        self.avg = [
                            (
                                self.average_aux[0]/self.average_aux[1],
                                float(cur_val[3]),
                                float(cur_val[3])
                            ),
                        ]
                        self.data.append(
                            (
                                str((next_dt+datetime.timedelta(hours=self.timezonediff)).date()),
                                str((next_dt+datetime.timedelta(hours=self.timezonediff)).time()),
                                *cur_val,
                            ),
                        )
                    else:
                        cur_val=('Disconnected',)*4
                        self.incrementdata=Increment('Disconnected')
                        self.average_aux=[None,None,None,None]
                        self.avg=[('Disconnected',)*3]
                        self.data.append(
                            (
                                str((next_dt+datetime.timedelta(hours=self.timezonediff)).date()),
                                str((next_dt+datetime.timedelta(hours=self.timezonediff)).time()),
                                *cur_val
                            ),
                        )
                    self.found_first[vo]=1
                else:
                        self.incrementdata.next(cur_val[3])
                        self.avg.append(
                            (
                                self.average_aux[0]/self.average_aux[1],
                                self.average_aux[2],
                                self.average_aux[3],
                            ),
                        )
                        self.data.append(
                            (
                                str((next_dt+datetime.timedelta(hours=self.timezonediff)).date()),
                                str((next_dt+datetime.timedelta(hours=self.timezonediff)).time()),
                                *cur_val                                               
                            ),
                        )
                next_dt=self.roundTime(next_dt,self.roundTo)                            
                while next_dt<dt and next_dt<self.tdate:
                    self.incrementdata.next('Disconnected')
                    self.avg.append(('Disconnected',)*3)
                    self.data.append(
                        (
                            str((next_dt+datetime.timedelta(hours=self.timezonediff)).date()),
                            str((next_dt+datetime.timedelta(hours=self.timezonediff)).time()),
                            *cur_val                                               
                        ),
                    )
                    next_dt=self.roundTime(next_dt,self.roundTo)
                cur_val=(vm,vo,vq,vv)
                self.average_aux[0]=self.average_aux[2]=self.average_aux[3]=vv
                self.average_aux[1]=1
            elif dt<=next_dt:
                cur_val=(vm,vo,vq,vv)
                if self.found_first[vo] == 0:
                    continue
                else:
                    if vv>self.average_aux[2]:
                        self.average_aux[2]=vv
                    elif vv<self.average_aux[3]:
                        self.average_aux[3]=vv
                    self.average_aux[0]+=vv
                    self.average_aux[1]+=1
        else:
            if self.found_first[self.uid] == 0:
                if cur_val:
                    self.incrementdata = Increment(cur_val[3])
                    self.average_aux = [
                        float(cur_val[3]),
                        1.0,
                        float(cur_val[3]),
                        float(cur_val[3])
                    ]
                    self.avg = [
                        (
                            self.average_aux[0]/self.average_aux[1],
                            float(cur_val[3]),
                            float(cur_val[3])
                        ),
                    ]
                    self.data.append(
                        (
                            str((next_dt+datetime.timedelta(hours=self.timezonediff)).date()),
                            str((next_dt+datetime.timedelta(hours=self.timezonediff)).time()),
                            *cur_val,
                        ),
                    )
                else:
                    cur_val=('Disconnected',)*4
                    self.incrementdata=Increment('Disconnected')
                    self.average_aux=[None,None,None,None]
                    self.avg=[('Disconnected',)*3]
                    self.data.append(
                        (
                            str((next_dt+datetime.timedelta(hours=self.timezonediff)).date()),
                            str((next_dt+datetime.timedelta(hours=self.timezonediff)).time()),
                            *cur_val
                        ),
                    )
                self.found_first[self.uid]=1
                next_dt=self.roundTime(next_dt,self.roundTo)
        while next_dt<=self.tdate:
                self.incrementdata.next('Disconnected')
                self.avg.append(('Disconnected',)*3)
                self.data.append(
                    (
                        str((next_dt+datetime.timedelta(hours=self.timezonediff)).date()),
                        str((next_dt+datetime.timedelta(hours=self.timezonediff)).time()),
                        *cur_val                                               
                    ),
                )
                next_dt=self.roundTime(next_dt,self.roundTo)

        if self.data:
            if self.data_type == 'Average':
                self.data = [i+j for i,j in zip(self.data,self.avg)]
            elif self.data_type == 'Consumption':
                self.data = [i+(k,) for i,k in zip(self.data,self.incrementdata.inc)]

    def analog_raw(self):
        for vm,vo,dt,ms,vq,vv in self.unparsed_data:
            dt+=datetime.timedelta(milliseconds=ms)
            if(dt>=self.fdate and dt<=self.tdate and  vo==self.uid):
                self.data.append(
                    (
                        str((dt+datetime.timedelta(hours=self.timezonediff)).date()),
                        str((dt+datetime.timedelta(hours=self.timezonediff)).time()),
                        vm,vo,vq,vv
                    ),
                )
    def event_selected(self):
        for dt,ms,em,eu in self.unparsed_data:
            dt+=datetime.timedelta(milliseconds=ms)
            eu = eu.replace('=',' / ')
            self.data.append(
                (
                    str((dt+datetime.timedelta(hours=self.timezonediff)).date()),
                    str((dt+datetime.timedelta(hours=self.timezonediff)).time()),
                    em,eu
                )
            )

    def roundTime(self,d1,roundTo):
        seconds=(d1-d1.min).seconds
        rounding=(seconds+roundTo/2)//roundTo*roundTo
        d2=d1+datetime.timedelta(0,rounding-seconds,-d1.microsecond)
        if(d2>d1):
            return d2
        else:
            return d2+datetime.timedelta(seconds=roundTo)
    def isListEmpty(self,inList):
        if isinstance(inList,list):
            return all(map(self.isListEmpty,inList))
        return False
class EventReport:
    def __init__(self,fdate,tdate,timezone,template_path=None):
        try:
            # self.str_time=datetime.datetime.now().strftime('%a %d %b %y %I-%M-%S%p')
            self.str_time=datetime.datetime.now().strftime('%y%m%d%H%M%S')
            if not os_path.isdir('./SignalLog'):
                mkdir('./SignalLog')
            self.file_path = './SignalLog/'+self.str_time+' event.xlsx'
            self.template_path = template_path #= "C:/Users/OISM/Desktop/SignalReportApp/Templates/template1.xlsx"
            self.fdate = fdate
            self.tdate = tdate
            self.timezonediff = (-1)*int(re.findall('GMT(.+):00',timezone)[0])
            self.fdate += datetime.timedelta(hours=self.timezonediff)
            self.tdate += datetime.timedelta(hours=self.timezonediff)
            self.timezonediff*=-1
            self.not_found=None
            self.data=[]
            if os_path.isfile('settings'):
                with open('settings','rb') as file:
                    self.customization = pickle.load(file)
            else:
                self.customization = Preferences.options_default

            signal = GetSignalDataEvent(self.fdate,self.tdate,**self.customization['MySQL'])
            self.unparsed_data = signal.result
            for dt,ms,oo,od,em,eq,eu in self.unparsed_data:
                dt+=datetime.timedelta(milliseconds=ms)
                eu = eu.replace('=',' / ')
                self.data.append(
                    (
                        str((dt+datetime.timedelta(hours=self.timezonediff)).date()),
                        str((dt+datetime.timedelta(hours=self.timezonediff)).time()),
                        oo,od,em,eq,eu,
                    )
                )
            if self.data:
                self.workbook = SingleSheetEvent(
                    self.data,
                    None,
                    None,
                    None,
                    self.customization,
                    self.template_path,
                    self.file_path,
                )
                self.result = 1
            else:
                self.result = 0
        except AccessDeniedError:
            self.result = -1
        except PermissionError:
            self.result = -2
        except ExcessiveDataError:
            self.result = -3
        except Exception as e:
            self.result = -4
            self.errormessage = str(e)
            print(e)

if __name__ == "__main__":

#     b=AnalogReport(datetime.datetime(2020,7,22,12,39),datetime.datetime(2020,7,22,12,55),'GMT+4:00',False,[
#       'MOSG / 33KV / H01_T40 TRF / MEASUREMENT / ACTIVE POWER(P)',
#       'MOSG / 33KV / H01_T40 TRF / MEASUREMENT / CURRENT IB',
#       'MOSG / 33KV / H01_T40 TRF / MEASUREMENT / VOLTAGE VRY',
#       'MOSG / 33KV / H01_T40 TRF / MEASUREMENT / FREQUENCY'
#   ],'All Measurements',15,1,'Average') #65,2,25 
#     b=ParseData(datetime.datetime(2019,7,22,12,39),datetime.datetime(2019,7,22,12,55),'GMT+4:00',True,[
#       'MOSG / 33KV / H01_T40 TRF / MEASUREMENT / ACTIVE POWER(P)',
#       'MOSG / 33KV / H01_T40 TRF / MEASUREMENT / CURRENT IB',
#       'MOSG / 33KV / H01_T40 TRF / MEASUREMENT / VOLTAGE VRY',
#       'MOSG / 33KV / H01_T40 TRF / MEASUREMENT / FREQUENCY'
#   ],'All Measurements') #65,2,25 
    # b=ParseData(datetime.datetime(2019,7,4,10,41),datetime.datetime(2019,7,4,11,0),'GMT+4:00',True,[
    #     'MOSG / 33KV / H12_13 BUS SEC / Q0 CB / POSITION'
    # ],'All Signals')
    # b=ParseData(datetime.datetime(2019,6,19,15,57),datetime.datetime(2019,6,19,16,0),'GMT+4:00',True,[
    #     'MOSG / 33KV / H06_T10 (LV) INC / GIS SIGNALS / GIS CB SPRING DISCHARGED',
    #     'MOSG / 33KV / H06_T10 (LV) INC / GIS SIGNALS / GIS DC SUPPLY FAIL',
    #     'MOSG / 33KV / H06_T10 (LV) INC / PROTECTION / 3-PH O_C AND E_F PROTECTION',
    # ],'All Signals')
    b = EventReport(datetime.datetime(2019,6,19,15,58),datetime.datetime(2019,6,19,16,00),'GMT+4:00','All Signals')
    print(b.result)
 